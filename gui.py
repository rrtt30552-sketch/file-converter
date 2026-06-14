"""
File Converter - 万能文件转换工具 (GUI 版)
双击即用，支持 PDF/Word/图片/Excel 常见格式互转。
"""

import os
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path


# ──────────────────────────────────────────────
# 转换引擎
# ──────────────────────────────────────────────

def convert_image(input_path, output_path, quality=95):
    from PIL import Image
    img = Image.open(input_path)
    ext = Path(output_path).suffix.lower().lstrip(".")
    if ext in ("jpg", "jpeg") and img.mode in ("RGBA", "P", "LA"):
        img = img.convert("RGB")
    kw = {}
    if ext in ("jpg", "jpeg"):
        kw = {"quality": quality, "optimize": True}
    elif ext == "webp":
        kw = {"quality": quality}
    elif ext == "png":
        kw = {"optimize": True}
    img.save(output_path, **kw)


def images_to_pdf(input_paths, output_path):
    from PIL import Image
    images = []
    for p in input_paths:
        img = Image.open(p)
        if img.mode in ("RGBA", "P", "LA"):
            img = img.convert("RGB")
        images.append(img)
    if not images:
        raise ValueError("没有有效的图片文件")
    first = images[0]
    rest = images[1:] if len(images) > 1 else []
    first.save(output_path, "PDF", save_all=True, append_images=rest, quality=95)


def pdf_to_images(input_path, output_dir, fmt="png", dpi=200):
    from pdf2image import convert_from_path
    os.makedirs(output_dir, exist_ok=True)
    images = convert_from_path(input_path, dpi=dpi, fmt=fmt)
    result = []
    stem = Path(input_path).stem
    for i, img in enumerate(images, 1):
        p = os.path.join(output_dir, f"{stem}_page{i}.{fmt}")
        img.save(p)
        result.append(p)
    return result


def merge_pdfs(input_paths, output_path):
    from pypdf import PdfReader, PdfWriter
    writer = PdfWriter()
    for p in input_paths:
        reader = PdfReader(p)
        for page in reader.pages:
            writer.add_page(page)
    with open(output_path, "wb") as f:
        writer.write(f)


def split_pdf(input_path, output_dir, pages=None):
    from pypdf import PdfReader, PdfWriter
    os.makedirs(output_dir, exist_ok=True)
    reader = PdfReader(input_path)
    total = len(reader.pages)
    if pages:
        nums = _parse_range(pages, total)
    else:
        nums = list(range(1, total + 1))
    result = []
    stem = Path(input_path).stem
    for n in nums:
        writer = PdfWriter()
        writer.add_page(reader.pages[n - 1])
        out = os.path.join(output_dir, f"{stem}_page{n}.pdf")
        with open(out, "wb") as f:
            writer.write(f)
        result.append(out)
    return result


def _parse_range(spec, total):
    nums = set()
    for part in spec.split(","):
        part = part.strip()
        if "-" in part:
            s, e = part.split("-", 1)
            nums.update(range(max(1, int(s)), min(total, int(e)) + 1))
        else:
            n = int(part)
            if 1 <= n <= total:
                nums.add(n)
    return sorted(nums)


def word_to_pdf(input_path, output_path):
    try:
        import subprocess
        subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf",
             "--outdir", str(Path(output_path).parent), input_path],
            capture_output=True, timeout=60
        )
        src = Path(input_path)
        generated = Path(output_path).parent / f"{src.stem}.pdf"
        if generated.exists() and str(generated) != output_path:
            generated.rename(output_path)
        return
    except Exception:
        pass

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
    from reportlab.lib.units import mm
    from docx import Document

    doc = Document(input_path)
    pdf = SimpleDocTemplate(output_path, pagesize=A4,
                            leftMargin=25*mm, rightMargin=25*mm,
                            topMargin=25*mm, bottomMargin=25*mm)
    styles = getSampleStyleSheet()
    story = []
    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            story.append(Spacer(1, 6*mm))
            continue
        sn = "Normal"
        if para.style and para.style.name:
            n = para.style.name.lower()
            if "heading 1" in n or "title" in n:
                sn = "Heading1"
            elif "heading 2" in n:
                sn = "Heading2"
            elif "heading 3" in n:
                sn = "Heading3"
        story.append(Paragraph(text, styles[sn]))
    pdf.build(story)


def excel_to_csv(input_path, output_path, sheet=None):
    import openpyxl, csv
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb.active
    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)
    wb.close()


def csv_to_excel(input_path, output_path):
    import csv, openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    with open(input_path, "r", encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            ws.append(row)
    wb.save(output_path)


def pdf_extract_text(input_path, output_path):
    from pypdf import PdfReader
    reader = PdfReader(input_path)
    text = "\n\n".join(page.extract_text() or "" for page in reader.pages)
    Path(output_path).write_text(text, encoding="utf-8")


# ──────────────────────────────────────────────
# 转换类型定义
# ──────────────────────────────────────────────

CONVERTERS = {
    "图片 → PDF": {
        "inputs": [("图片文件", "*.png;*.jpg;*.jpeg;*.webp;*.bmp;*.tiff;*.gif")],
        "output_ext": ".pdf",
        "multi_input": True,
        "func": lambda ins, out, _: images_to_pdf(ins, out),
    },
    "PDF → 图片": {
        "inputs": [("PDF 文件", "*.pdf")],
        "output_ext": None,  # 输出目录
        "multi_input": False,
        "func": lambda i, o, _: pdf_to_images(i, o),
        "output_is_dir": True,
    },
    "图片格式转换": {
        "inputs": [("图片文件", "*.png;*.jpg;*.jpeg;*.webp;*.bmp;*.tiff;*.gif")],
        "output_ext": ".png",
        "multi_input": False,
        "func": lambda i, o, _: convert_image(i, o),
        "ext_choices": [".png", ".jpg", ".webp", ".bmp", ".tiff"],
    },
    "Word → PDF": {
        "inputs": [("Word 文件", "*.docx")],
        "output_ext": ".pdf",
        "multi_input": False,
        "func": lambda i, o, _: word_to_pdf(i, o),
    },
    "PDF → 文本": {
        "inputs": [("PDF 文件", "*.pdf")],
        "output_ext": ".txt",
        "multi_input": False,
        "func": lambda i, o, _: pdf_extract_text(i, o),
    },
    "合并 PDF": {
        "inputs": [("PDF 文件", "*.pdf")],
        "output_ext": ".pdf",
        "multi_input": True,
        "func": lambda ins, out, _: merge_pdfs(ins, out),
    },
    "拆分 PDF": {
        "inputs": [("PDF 文件", "*.pdf")],
        "output_ext": None,
        "multi_input": False,
        "func": lambda i, o, _: split_pdf(i, o),
        "output_is_dir": True,
    },
    "Excel → CSV": {
        "inputs": [("Excel 文件", "*.xlsx;*.xls")],
        "output_ext": ".csv",
        "multi_input": False,
        "func": lambda i, o, _: excel_to_csv(i, o),
    },
    "CSV → Excel": {
        "inputs": [("CSV 文件", "*.csv")],
        "output_ext": ".xlsx",
        "multi_input": False,
        "func": lambda i, o, _: csv_to_excel(i, o),
    },
}


# ──────────────────────────────────────────────
# GUI
# ──────────────────────────────────────────────

class FileConverterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🔄 万能文件转换工具")
        self.root.geometry("680x520")
        self.root.resizable(False, False)

        # 配色
        self.bg = "#f5f5f5"
        self.accent = "#4a90d9"
        self.success = "#4caf50"
        self.root.configure(bg=self.bg)

        self.selected_files = []
        self._build_ui()

    def _build_ui(self):
        # ── 标题 ──
        header = tk.Frame(self.root, bg=self.accent, height=60)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="🔄 万能文件转换工具",
                 font=("Microsoft YaHei UI", 18, "bold"),
                 fg="white", bg=self.accent).pack(expand=True)

        # ── 主体 ──
        main = tk.Frame(self.root, bg=self.bg, padx=30, pady=15)
        main.pack(fill="both", expand=True)

        # 转换类型
        tk.Label(main, text="选择转换类型：",
                 font=("Microsoft YaHei UI", 11), bg=self.bg).pack(anchor="w")

        self.type_var = tk.StringVar(value="图片 → PDF")
        type_combo = ttk.Combobox(main, textvariable=self.type_var,
                                  values=list(CONVERTERS.keys()),
                                  state="readonly", width=30,
                                  font=("Microsoft YaHei UI", 10))
        type_combo.pack(anchor="w", pady=(2, 12))
        type_combo.bind("<<ComboboxSelected>>", self._on_type_change)

        # 文件选择区
        file_frame = tk.Frame(main, bg=self.bg)
        file_frame.pack(fill="x")

        self.file_label = tk.Label(file_frame, text="未选择文件",
                                   font=("Microsoft YaHei UI", 10),
                                   fg="#888", bg=self.bg, anchor="w")
        self.file_label.pack(side="left", fill="x", expand=True)

        btn_style = {"font": ("Microsoft YaHei UI", 10), "cursor": "hand2",
                     "relief": "flat", "padx": 12, "pady": 4}

        self.select_btn = tk.Button(file_frame, text="📁 选择文件",
                                    bg=self.accent, fg="white",
                                    command=self._select_files, **btn_style)
        self.select_btn.pack(side="right")

        # 输出格式选择（图片格式转换时显示）
        self.ext_frame = tk.Frame(main, bg=self.bg)
        tk.Label(self.ext_frame, text="输出格式：",
                 font=("Microsoft YaHei UI", 10), bg=self.bg).pack(side="left")
        self.ext_var = tk.StringVar(value=".png")
        self.ext_combo = ttk.Combobox(self.ext_frame, textvariable=self.ext_var,
                                      state="readonly", width=10,
                                      font=("Microsoft YaHei UI", 10))
        self.ext_combo.pack(side="left", padx=(5, 0))

        # 文件列表
        list_frame = tk.Frame(main, bg="white", relief="solid", bd=1)
        list_frame.pack(fill="both", expand=True, pady=(12, 12))

        self.file_listbox = tk.Listbox(list_frame, font=("Consolas", 9),
                                       selectmode="extended", bd=0,
                                       highlightthickness=0)
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical",
                                  command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # 转换按钮
        self.convert_btn = tk.Button(main, text="⚡ 开始转换",
                                     font=("Microsoft YaHei UI", 13, "bold"),
                                     bg=self.success, fg="white",
                                     cursor="hand2", relief="flat",
                                     padx=30, pady=8,
                                     command=self._start_convert)
        self.convert_btn.pack(pady=(0, 8))

        # 状态栏
        self.status_var = tk.StringVar(value="就绪")
        self.status_bar = tk.Label(main, textvariable=self.status_var,
                                   font=("Microsoft YaHei UI", 9),
                                   fg="#666", bg=self.bg, anchor="w")
        self.status_bar.pack(fill="x")

        # 进度条
        self.progress = ttk.Progressbar(main, mode="determinate", length=400)

        self._on_type_change()

    def _on_type_change(self, event=None):
        conv = CONVERTERS[self.type_var.get()]
        if conv.get("ext_choices"):
            self.ext_frame.pack(fill="x", pady=(0, 5))
            self.ext_combo["values"] = conv["ext_choices"]
            self.ext_var.set(conv["ext_choices"][0])
        else:
            self.ext_frame.pack_forget()
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self.file_label.config(text="未选择文件", fg="#888")

    def _select_files(self):
        conv = CONVERTERS[self.type_var.get()]
        filetypes = conv["inputs"]

        if conv["multi_input"]:
            files = filedialog.askopenfilenames(filetypes=filetypes + [("所有文件", "*.*")])
        else:
            f = filedialog.askopenfilename(filetypes=filetypes + [("所有文件", "*.*")])
            files = [f] if f else []

        if files:
            self.selected_files = list(files)
            self.file_listbox.delete(0, tk.END)
            for f in self.selected_files:
                self.file_listbox.insert(tk.END, f)
            count = len(self.selected_files)
            self.file_label.config(
                text=f"已选择 {count} 个文件" if count > 1 else self.selected_files[0],
                fg="#333")

    def _start_convert(self):
        if not self.selected_files:
            messagebox.showwarning("提示", "请先选择文件！")
            return

        conv = CONVERTERS[self.type_var.get()]

        # 确定输出路径
        if conv.get("output_is_dir"):
            out = filedialog.askdirectory(title="选择输出目录")
            if not out:
                return
        else:
            ext = self.ext_var.get() if conv.get("ext_choices") else conv["output_ext"]
            default_name = Path(self.selected_files[0]).stem + ext
            out = filedialog.asksaveasfilename(
                defaultextension=ext,
                initialfile=default_name,
                filetypes=[("输出文件", f"*{ext}"), ("所有文件", "*.*")]
            )
            if not out:
                return

        # 禁用按钮，显示进度
        self.convert_btn.config(state="disabled", text="转换中...")
        self.progress.pack(fill="x", pady=(5, 0))
        self.progress["value"] = 0

        # 在线程中执行转换
        threading.Thread(
            target=self._do_convert,
            args=(conv, out),
            daemon=True
        ).start()

    def _do_convert(self, conv, output):
        try:
            ins = self.selected_files
            ext = self.ext_var.get() if conv.get("ext_choices") else None

            if conv.get("output_is_dir"):
                # 输出是目录（PDF→图片、拆分PDF）
                conv["func"](ins[0], output, None)
            elif len(ins) > 1 and conv["multi_input"]:
                conv["func"](ins, output, ext)
            else:
                conv["func"](ins[0], output, ext)

            self.root.after(0, self._convert_done, True, output)
        except Exception as e:
            self.root.after(0, self._convert_done, False, str(e))

    def _convert_done(self, success, info):
        self.progress.pack_forget()
        self.convert_btn.config(state="normal", text="⚡ 开始转换")

        if success:
            self.status_var.set(f"✅ 转换完成！输出: {info}")
            messagebox.showinfo("完成", f"转换成功！\n\n输出位置：{info}")
        else:
            self.status_var.set(f"❌ 转换失败")
            messagebox.showerror("失败", f"转换出错：{info}")


def main():
    root = tk.Tk()

    # 设置样式
    style = ttk.Style()
    style.theme_use("clam")

    app = FileConverterApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
