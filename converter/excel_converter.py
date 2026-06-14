"""Excel / CSV conversion utilities."""

from pathlib import Path


def excel_to_csv(input_path: str, output_path: str, sheet_name: str | None = None) -> str:
    """Convert an Excel file (.xlsx) to CSV.

    Args:
        input_path: Source .xlsx file path.
        output_path: Target .csv file path.
        sheet_name: Sheet to convert. None = first sheet.

    Returns:
        Absolute path of the output CSV.
    """
    import openpyxl

    src = Path(input_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(str(src), read_only=True, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    import csv
    with open(dst, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)

    wb.close()
    return str(dst.resolve())


def csv_to_excel(input_path: str, output_path: str) -> str:
    """Convert a CSV file to Excel (.xlsx).

    Args:
        input_path: Source .csv file path.
        output_path: Target .xlsx file path.

    Returns:
        Absolute path of the output Excel file.
    """
    import csv
    import openpyxl

    src = Path(input_path)
    dst = Path(output_path)
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active

    with open(src, "r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)

    wb.save(str(dst))
    return str(dst.resolve())
